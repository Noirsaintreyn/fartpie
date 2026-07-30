"""
Builds an .xlsx report (openable in Numbers) from backtest_summary_fixed.csv,
split into 1H and 4H sheets, with NQ/ES combined roll-ups computed via
formula (weighted by touch/break counts, not a naive average of rates).
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SUMMARY_CSV = 'backtest_summary_fixed.csv'
OUT_XLSX = 'backtest_results_1h_4h.xlsx'

FONT_NAME = 'Arial'
HEADER_FILL = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
HEADER_FONT = Font(name=FONT_NAME, bold=True, color='FFFFFF')
BOLD = Font(name=FONT_NAME, bold=True)
NORMAL = Font(name=FONT_NAME)

COLUMNS = [
    ('Instrument', 14),
    ('Method', 18),
    ('N Levels Generated', 18),
    ('N Touched', 12),
    ('Touch Rate', 12),
    ('Support Accuracy', 16),
    ('Break Rate', 12),
    ('False Breakout Recovery', 20),
    ('Avg Hold (bars)', 14),
    ('N Bounced', 12),
    ('N Broken', 12),
    ('N Recovered', 12),
]

PCT_COLS = {5, 6, 7, 8}   # 1-indexed: Touch Rate, Support Accuracy, Break Rate, False Breakout Recovery
NUM_COLS = {3, 4, 9, 10, 11, 12}


def write_sheet(ws, df_tf, methods):
    for j, (name, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=j, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(j)].width = width
    ws.freeze_panes = 'A2'

    row = 2
    for method in methods:
        sub = df_tf[df_tf['method'] == method]
        nq = sub[sub['instrument'] == 'NQ']
        es = sub[sub['instrument'] == 'ES']
        rows_written = []
        for inst, r in [('NQ', nq), ('ES', es)]:
            if r.empty:
                continue
            r = r.iloc[0]
            ws.cell(row=row, column=1, value=inst).font = NORMAL
            ws.cell(row=row, column=2, value=method).font = NORMAL
            ws.cell(row=row, column=3, value=int(r['n_levels_generated']))
            ws.cell(row=row, column=4, value=int(r['n_touched']))
            ws.cell(row=row, column=5, value=float(r['touch_rate']))
            ws.cell(row=row, column=6, value=float(r['support_accuracy']))
            ws.cell(row=row, column=7, value=float(r['break_rate']))
            ws.cell(row=row, column=8, value=float(r['false_breakout_recovery']))
            ws.cell(row=row, column=9, value=float(r['avg_hold_duration_bars']))
            # helper counts, derived via formula from the rate + denominator cells on this row
            ws.cell(row=row, column=10, value=f'=ROUND(F{row}*D{row},0)')       # N Bounced = support_accuracy * n_touched
            ws.cell(row=row, column=11, value=f'=ROUND(G{row}*D{row},0)')       # N Broken  = break_rate * n_touched
            ws.cell(row=row, column=12, value=f'=ROUND(H{row}*K{row},0)')       # N Recovered = false_breakout_recovery * n_broken
            rows_written.append(row)
            row += 1

        if len(rows_written) == 2:
            r1, r2 = rows_written
            ws.cell(row=row, column=1, value='Combined').font = BOLD
            ws.cell(row=row, column=2, value=method).font = BOLD
            ws.cell(row=row, column=3, value=f'=SUM(C{r1}:C{r2})').font = BOLD
            ws.cell(row=row, column=4, value=f'=SUM(D{r1}:D{r2})').font = BOLD
            ws.cell(row=row, column=5, value=f'=SUM(D{r1}:D{r2})/SUM(C{r1}:C{r2})').font = BOLD
            # weighted by n_touched (not a naive average of the two rates)
            ws.cell(row=row, column=6, value=f'=SUMPRODUCT(F{r1}:F{r2},D{r1}:D{r2})/SUM(D{r1}:D{r2})').font = BOLD
            ws.cell(row=row, column=7, value=f'=SUMPRODUCT(G{r1}:G{r2},D{r1}:D{r2})/SUM(D{r1}:D{r2})').font = BOLD
            # weighted by n_broken for recovery specifically
            ws.cell(row=row, column=8, value=f'=SUM(L{r1}:L{r2})/SUM(K{r1}:K{r2})').font = BOLD
            ws.cell(row=row, column=9, value=f'=SUMPRODUCT(I{r1}:I{r2},D{r1}:D{r2})/SUM(D{r1}:D{r2})').font = BOLD
            ws.cell(row=row, column=10, value=f'=SUM(J{r1}:J{r2})').font = BOLD
            ws.cell(row=row, column=11, value=f'=SUM(K{r1}:K{r2})').font = BOLD
            ws.cell(row=row, column=12, value=f'=SUM(L{r1}:L{r2})').font = BOLD
            row += 1

    for r in range(2, row):
        for c in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=r, column=c)
            if c in PCT_COLS:
                cell.number_format = '0.0%'
            elif c == 9:
                cell.number_format = '0.00'
            elif c in NUM_COLS:
                cell.number_format = '#,##0'
    return row


def write_notes_sheet(ws, method_rank):
    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['B'].width = 90
    lines = [
        ('Backtest methodology & notes', ''),
        ('', ''),
        ('Data source', 'User-provided CSVs: 1H_NQ.csv, 4H_NQ.csv, 1H_ES.csv, 4H_ES.csv (continuous front-month futures, 2014-01-02 to 2026-01-30)'),
        ('Detection window (lookback)', '150 bars'),
        ('Forward evaluation horizon', '40 bars'),
        ('Step between windows', '20 bars (walk-forward, no overlap-driven leakage; each window\'s levels are generated only from bars strictly before it)'),
        ('Bounce threshold', '1.0x ATR (14-period, computed from the detection window only) favorable move required to count as a bounce'),
        ('Break threshold', '0.5x ATR adverse close-through'),
        ('Break confirmation', '2 consecutive closes beyond the break level required (fixes an earlier bug where a single noisy close instantly killed the bounce case before it had a chance to develop)'),
        ('Reaction window', '10 bars after touch to resolve bounce vs. break'),
        ('Recovery window', '10 bars after a confirmed break to check whether price reclaims the level (false breakout)'),
        ('', ''),
        ('Metric definitions', ''),
        ('Touch Rate', 'Share of generated levels that price actually traded through within the horizon'),
        ('Support Accuracy', 'Of touched levels, share that bounced (favorable move >= bounce threshold) before a confirmed break'),
        ('Break Rate', 'Of touched levels, share that were decisively broken (2+ consecutive closes past the break threshold)'),
        ('False Breakout Recovery', 'Of broken levels, share where price reclaimed the level within the recovery window'),
        ('Avg Hold (bars)', 'Average bars from touch to confirmed break, censored at the reaction window if never broken'),
        ('', ''),
        ('Known limitation', 'Data is raw stitched contract-month futures (not back-adjusted), so small price gaps exist at each quarterly rollover. This affects all methods equally and does not bias relative comparisons.'),
        ('Known limitation', 'GMM/Wavelet/HMM-Levels have no persistent training - they refit fresh on every window, same as HDBSCAN/OPTICS/KDE, so there is no train/test leakage risk for these methods.'),
    ]
    for i, (a, b) in enumerate(lines, start=1):
        is_header = i == 1
        is_section = a in ('Metric definitions',)
        cell_a = ws.cell(row=i, column=1, value=a)
        cell_b = ws.cell(row=i, column=2, value=b)
        cell_a.font = Font(name=FONT_NAME, bold=True, size=14) if is_header else (BOLD if (b == '' and a) or is_section else BOLD)
        cell_b.font = NORMAL
        cell_a.alignment = Alignment(vertical='top', wrap_text=True)
        cell_b.alignment = Alignment(vertical='top', wrap_text=True)


def main():
    df = pd.read_csv(SUMMARY_CSV)
    wb = Workbook()
    wb.remove(wb.active)

    for tf, sheet_name in [('1h', '1H Results'), ('4h', '4H Results')]:
        df_tf = df[df['timeframe'] == tf].copy()
        # rank methods within this timeframe by combined (NQ+ES) support accuracy for display order
        method_order = (
            df_tf.groupby('method')
            .apply(lambda g: (g['support_accuracy'] * g['n_touched']).sum() / g['n_touched'].sum())
            .sort_values(ascending=False)
            .index.tolist()
        )
        ws = wb.create_sheet(sheet_name)
        write_sheet(ws, df_tf, method_order)

    notes_ws = wb.create_sheet('Methodology & Notes')
    write_notes_sheet(notes_ws, None)

    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX}")


if __name__ == '__main__':
    main()
